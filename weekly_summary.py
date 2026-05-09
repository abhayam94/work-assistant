"""
weekly_summary.py — Friday end-of-week summary.
Run: python weekly_summary.py

Shows:
  - Hours logged this week
  - Tickets closed / moved to review
  - Tickets still open and overdue
  - Jira comments received this week (on your tickets)
"""

import textwrap
import requests
from datetime import date, timedelta
from colorama import init, Fore, Style
from openpyxl import load_workbook
from jira_client import get_my_open_tickets, get_resolved_tickets_since, get_ticket_url
from logger import log_start, log_end, log_error
from config import LOCAL_TIMESHEET_PATH, ALL_END_STATUSES, ALL_CLOSE_STATUSES

init(autoreset=True)


def _excel_to_date(val):
    from datetime import datetime as dt_type, date as date_type
    if isinstance(val, dt_type):
        return val.date()
    if isinstance(val, date_type):
        return val
    if isinstance(val, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(val))
    return None


def print_divider(char="─", width=70):
    print(Fore.WHITE + Style.DIM + char * width)


def get_week_bounds():
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    friday = monday + timedelta(days=4)
    return monday, friday


def get_hours_this_week(monday: date, friday: date):
    """Read hours from In-Out sheet for Mon–Fri this week."""
    try:
        wb = load_workbook(LOCAL_TIMESHEET_PATH, data_only=True)
        ws = wb["In-Out"]
        total = 0.0
        days_logged = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = _excel_to_date(row[0])
            if d and monday <= d <= friday:
                hours = row[2] or 0
                total += float(hours)
                days_logged.append((d, row[1] or "", float(hours)))
        return total, days_logged
    except Exception as e:
        return 0.0, []


def get_tickets_summary():
    """From Tickets sheet, find what moved this week."""
    monday, friday = get_week_bounds()
    try:
        wb = load_workbook(LOCAL_TIMESHEET_PATH, data_only=True)
        ws = wb["Tickets"]
        closed_this_week = []
        reviewed_this_week = []
        still_open = []
        overdue = []
        today = date.today()

        for row in ws.iter_rows(min_row=2, values_only=True):
            ticket = str(row[1]).strip() if row[1] else ""
            if not ticket or ticket == "None":
                continue
            status   = row[2] or ""
            started  = _excel_to_date(row[3])
            end_d    = _excel_to_date(row[4])
            close_d  = _excel_to_date(row[5])
            due_d    = _excel_to_date(row[8])
            report   = row[10] or ""

            entry = {"ticket": ticket, "status": status, "report": report,
                     "url": get_ticket_url(ticket)}

            if close_d and monday <= close_d <= friday:
                closed_this_week.append(entry)
            elif end_d and monday <= end_d <= friday:
                reviewed_this_week.append(entry)
            elif status.lower() not in [s.lower() for s in ALL_CLOSE_STATUSES]:
                still_open.append(entry)
                if due_d and due_d < today:
                    overdue.append({**entry, "due": due_d})

        return closed_this_week, reviewed_this_week, still_open, overdue
    except Exception as e:
        return [], [], [], []


def get_recent_comments():
    """Fetch recent Jira comments on your open tickets from this week."""
    from requests.auth import HTTPBasicAuth
    from config import JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN
    monday, _ = get_week_bounds()
    auth = HTTPBasicAuth(JIRA_EMAIL, JIRA_API_TOKEN)
    headers = {"Accept": "application/json"}

    try:
        open_tickets = get_my_open_tickets()
        comments_found = []
        for issue in open_tickets[:10]:  # limit API calls
            key = issue["key"]
            url = f"{JIRA_BASE_URL}/rest/api/3/issue/{key}/comment"
            resp = requests.get(url, auth=auth, headers=headers, timeout=10)
            if resp.status_code != 200:
                continue
            comments = resp.json().get("comments", [])
            for c in comments[-3:]:  # last 3 comments per ticket
                created_str = c.get("created", "")[:10]
                try:
                    created = date.fromisoformat(created_str)
                except Exception:
                    continue
                if created >= monday:
                    author = c.get("author", {}).get("displayName", "Unknown")
                    # Extract plain text from ADF
                    body = c.get("body", {})
                    texts = []
                    def walk(node):
                        if isinstance(node, dict):
                            if node.get("type") == "text":
                                texts.append(node.get("text", ""))
                            for child in node.get("content", []):
                                walk(child)
                    walk(body)
                    text = " ".join(t.strip() for t in texts if t.strip())[:200]
                    comments_found.append({
                        "ticket": key,
                        "author": author,
                        "date": created_str,
                        "text": text
                    })
        return comments_found
    except Exception:
        return []


def run():
    log_start("weekly_summary")
    monday, friday = get_week_bounds()
    today = date.today()

    print()
    print(Fore.CYAN + Style.BRIGHT + "  WORK ASSISTANT — Weekly Summary")
    print(Fore.WHITE + Style.DIM +
          f"  Week of {monday.strftime('%d %b')} – {friday.strftime('%d %b %Y')}")
    print_divider("═")

    # ── Hours ──────────────────────────────────────────────────────
    print(Fore.CYAN + Style.BRIGHT + "\n  ⏱  Hours This Week\n")
    total_hours, days = get_hours_this_week(monday, friday)
    if days:
        for d, task, h in days:
            print(f"  {Fore.YELLOW}{d.strftime('%a %d %b'):<14}"
                  f"{Fore.WHITE}{str(h)+'h':<8}{task[:50]}")
        print_divider()
        print(f"  {Fore.GREEN + Style.BRIGHT}Total: {total_hours}h  "
              f"{Fore.WHITE + Style.DIM}({len(days)} days logged)")
    else:
        print(Fore.YELLOW + "  No hours logged this week yet.")

    # ── Tickets ────────────────────────────────────────────────────
    closed, reviewed, open_tickets, overdue = get_tickets_summary()

    print(Fore.CYAN + Style.BRIGHT + "\n  ✅  Closed This Week\n")
    if closed:
        for t in closed:
            print(f"  {Fore.GREEN}{t['ticket']:<14}{Fore.WHITE}{t['report']:<14}{t['status']}")
    else:
        print(Fore.WHITE + Style.DIM + "  None closed this week.")

    print(Fore.CYAN + Style.BRIGHT + "\n  🔍  Moved to Review This Week\n")
    if reviewed:
        for t in reviewed:
            print(f"  {Fore.CYAN}{t['ticket']:<14}{Fore.WHITE}{t['report']:<14}{t['status']}")
    else:
        print(Fore.WHITE + Style.DIM + "  None moved to review this week.")

    if overdue:
        print(Fore.RED + Style.BRIGHT + "\n  🚨  Overdue Open Tickets\n")
        for t in overdue:
            print(f"  {Fore.RED}{t['ticket']:<14}{Fore.WHITE}{t['report']:<14}"
                  f"{Fore.RED}Due: {t['due'].strftime('%d-%b-%Y')}")

    print(Fore.CYAN + Style.BRIGHT + f"\n  📋  Still Open ({len(open_tickets)} tickets)\n")
    for t in open_tickets[:10]:
        print(f"  {Fore.WHITE}{t['ticket']:<14}{Fore.YELLOW}{t['status']:<25}{t['report']}")
    if len(open_tickets) > 10:
        print(Fore.WHITE + Style.DIM + f"  ... and {len(open_tickets)-10} more")

    # ── Comments ───────────────────────────────────────────────────
    print(Fore.CYAN + Style.BRIGHT + "\n  💬  New Jira Comments This Week\n")
    try:
        comments = get_recent_comments()
        if comments:
            for c in comments:
                print(f"  {Fore.YELLOW}{c['ticket']:<12}"
                      f"{Fore.CYAN}{c['author']:<22}"
                      f"{Fore.WHITE + Style.DIM}{c['date']}")
                wrapped = textwrap.fill(c["text"], width=65,
                                        initial_indent="    ",
                                        subsequent_indent="    ")
                print(Fore.WHITE + wrapped)
                print()
        else:
            print(Fore.WHITE + Style.DIM + "  No new comments this week.")
    except Exception as e:
        log_error("get_recent_comments", e)
        print(Fore.YELLOW + f"  Could not fetch comments: {e}")

    print()
    print_divider("═")
    print(Fore.GREEN + Style.BRIGHT + "  Good work this week! 🎯\n")
    log_end("weekly_summary", success=True)


if __name__ == "__main__":
    run()
