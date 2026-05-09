"""
migrate_tickets_sheet.py — ONE TIME ONLY.
Migrates the old Tickets sheet structure to the new expanded format.

Old: Sl.No. | Ticket | Status | Started On | Completed On | Report | Comments
New: Sl.No. | Ticket | Status | Started On | End Date | Close Date | Days Open | Cycle Time | Due Date | Created Date | Report | Comments

Run BEFORE eod_updater.py for the first time.
Creates a backup: Timesheet_PRE_MIGRATION.xlsx
"""

import shutil
import copy
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Style
from config import LOCAL_TIMESHEET_PATH

init(autoreset=True)

NEW_HEADERS = [
    "Sl.No.", "Ticket", "Status", "Started On", "End Date",
    "Close Date", "Days Open", "Cycle Time", "Due Date",
    "Created Date", "Report", "Comments"
]

HEADER_FILL = PatternFill("solid", start_color="D9E1F2")
HEADER_FONT = Font(bold=True)

COL_WIDTHS = [6, 14, 22, 13, 13, 13, 10, 11, 13, 13, 14, 35]


def run():
    print()
    print(Fore.CYAN + Style.BRIGHT + "  TICKETS SHEET MIGRATION")
    print(Fore.WHITE + Style.DIM + "  One-time column restructure")
    print(Fore.WHITE + "─" * 60)

    # --- Backup first ---
    backup_path = LOCAL_TIMESHEET_PATH.replace(".xlsx", "_PRE_MIGRATION.xlsx")
    shutil.copy2(LOCAL_TIMESHEET_PATH, backup_path)
    print(Fore.GREEN + f"\n  ✓ Backup saved: {backup_path}")

    wb = load_workbook(LOCAL_TIMESHEET_PATH)

    if "Tickets" not in wb.sheetnames:
        print(Fore.RED + "  ✗ No 'Tickets' sheet found.")
        return

    ws_old = wb["Tickets"]

    # Read all existing data
    rows = list(ws_old.iter_rows(values_only=True))
    if not rows:
        print(Fore.RED + "  ✗ Tickets sheet is empty.")
        return

    # Map old columns by header name (case-insensitive)
    old_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    print(Fore.WHITE + f"\n  Old headers found: {[r for r in rows[0] if r]}")

    def col_idx(name):
        """Return 0-based index of old column by name."""
        for i, h in enumerate(old_headers):
            if name.lower() in h:
                return i
        return None

    idx_slno     = col_idx("sl")
    idx_ticket   = col_idx("ticket")
    idx_status   = col_idx("status")
    idx_started  = col_idx("started")
    idx_completed = col_idx("completed")
    idx_report   = col_idx("report")
    idx_comments = col_idx("comment")

    # Delete old sheet and recreate
    del wb["Tickets"]
    ws = wb.create_sheet("Tickets")

    # Write new headers
    for col, (header, width) in enumerate(zip(NEW_HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data rows
    migrated = 0
    for row_data in rows[1:]:
        if not any(row_data):
            continue

        def get(idx):
            if idx is None:
                return None
            try:
                return row_data[idx]
            except IndexError:
                return None

        sl       = get(idx_slno)
        ticket   = get(idx_ticket)
        status   = get(idx_status)
        started  = get(idx_started)
        end_date = get(idx_completed)   # Completed On → End Date
        report   = get(idx_report)
        comments = get(idx_comments)

        if not ticket:
            continue

        next_row = ws.max_row + 1

        ws.cell(next_row, 1).value = sl
        ws.cell(next_row, 2).value = str(ticket).strip()
        ws.cell(next_row, 3).value = status

        # Date cells — preserve format
        for col, val in [(4, started), (5, end_date)]:
            cell = ws.cell(next_row, col)
            cell.value = val
            if val:
                cell.number_format = "DD-MMM-YY"

        # Cols 6–10 (Close Date, Days Open, Cycle Time, Due Date, Created Date) — blank for now
        for col in range(6, 11):
            ws.cell(next_row, col).value = None

        ws.cell(next_row, 11).value = report
        ws.cell(next_row, 12).value = comments

        migrated += 1

    wb.save(LOCAL_TIMESHEET_PATH)

    print(Fore.GREEN + f"  ✓ Migration complete — {migrated} rows migrated")
    print(Fore.WHITE + Style.DIM + "\n  New column order:")
    for i, h in enumerate(NEW_HEADERS, 1):
        print(Fore.WHITE + Style.DIM + f"    Col {i}: {h}")

    print()
    print(Fore.YELLOW + "  Next step: run python eod_updater.py")
    print(Fore.WHITE + Style.DIM + f"  Backup kept at: {backup_path}\n")


if __name__ == "__main__":
    run()
