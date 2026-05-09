"""
timesheet_updater.py — Updates local Excel timesheet (In-Out + Tickets sheets)
Called by eod_updater.py after confirming hours with user.
"""

import os
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from config import (
    LOCAL_TIMESHEET_PATH,
    ALL_STARTED_STATUSES, ALL_END_STATUSES, ALL_CLOSE_STATUSES
)
from jira_client import (
    get_my_open_tickets, get_resolved_tickets_since, get_ticket_changelog,
    get_ticket_url, extract_status_dates, get_report_name, format_ticket_for_digest
)


# --- HELPERS ---

def _excel_date(d: date) -> int:
    """Convert Python date to Excel serial number."""
    from datetime import date as dt
    delta = d - dt(1899, 12, 30)
    return delta.days


def _time_to_decimal(time_str: str) -> float:
    """Convert 'HH:MM' string to Excel time decimal (fraction of day)."""
    h, m = map(int, time_str.split(":"))
    return (h * 60 + m) / 1440


def _get_or_create_month_sheet(wb, today: date):
    """Return the sheet for current month, creating if not exists."""
    month_name = today.strftime("%b %y")  # e.g. 'May 26'
    if month_name in wb.sheetnames:
        return wb[month_name]
    ws = wb.create_sheet(month_name)
    ws.append(["Date", "Task", "Hours"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)
    # Pre-populate all days for new month
    from format_sheets import format_monthly_sheet
    format_monthly_sheet(ws, today.year, today.month)
    return ws


def _find_inout_row_for_today(ws_inout, today_serial: int):
    """Return row index in In-Out sheet matching today's date, or None.
    Handles Excel serial integers, Python date, and datetime objects."""
    from datetime import datetime as dt_type, date as date_type
    today = date.today()
    for row in ws_inout.iter_rows(min_row=2, values_only=False):
        cell = row[0]
        val = cell.value
        if val is None:
            continue
        if isinstance(val, (int, float)) and int(val) == today_serial:
            return cell.row
        if isinstance(val, dt_type) and val.date() == today:
            return cell.row
        if isinstance(val, date_type) and not isinstance(val, dt_type) and val == today:
            return cell.row
    return None


def _get_tickets_sheet(wb):
    if "Tickets" not in wb.sheetnames:
        ws = wb.create_sheet("Tickets")
        headers = ["Sl.No.", "Ticket", "Status", "Started On", "End Date",
                   "Close Date", "Days Open", "Cycle Time", "Due Date",
                   "Created Date", "Report", "Comments"]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", start_color="D9E1F2")
        return ws
    return wb["Tickets"]


def _find_ticket_row(ws_tickets, ticket_key: str):
    """Return row index for existing ticket key, or None."""
    for row in ws_tickets.iter_rows(min_row=2, values_only=False):
        cell = row[1]  # Column B = Ticket
        val = cell.value
        if val is None:
            continue
        # Handle hyperlinked cell — value may be the display text
        display = str(val).strip()
        if display == ticket_key:
            return cell.row
    return None


def _write_date_cell(cell, d):
    """Write a date into a cell with date formatting."""
    if d is None:
        return
    if isinstance(d, date):
        cell.value = _excel_date(d)
        cell.number_format = "DD-MMM-YY"
    else:
        cell.value = d


# --- MAIN UPDATE FUNCTIONS ---

def update_inout_sheet(sign_in: str, sign_out: str, task_summary: str, hours: float):
    """
    Add or update today's row in the In-Out sheet and current month sheet.
    sign_in / sign_out: 'HH:MM' strings
    task_summary: brief description of today's work
    hours: float
    """
    wb = load_workbook(LOCAL_TIMESHEET_PATH)
    today = date.today()
    today_serial = _excel_date(today)

    # --- In-Out sheet ---
    ws_io = wb["In-Out"]
    existing_row = _find_inout_row_for_today(ws_io, today_serial)

    if existing_row:
        ws_io.cell(existing_row, 2).value = task_summary
        ws_io.cell(existing_row, 3).value = hours
        ws_io.cell(existing_row, 4).value = _time_to_decimal(sign_in)
        ws_io.cell(existing_row, 5).value = _time_to_decimal(sign_out)
    else:
        row = [today_serial, task_summary, hours,
               _time_to_decimal(sign_in), _time_to_decimal(sign_out)]
        ws_io.append(row)
        last_row = ws_io.max_row
        ws_io.cell(last_row, 1).number_format = "DD-MMM-YY"
        ws_io.cell(last_row, 4).number_format = "HH:MM"
        ws_io.cell(last_row, 5).number_format = "HH:MM"

    # --- Monthly sheet ---
    ws_month = _get_or_create_month_sheet(wb, today)
    # Check if today already has an entry
    from datetime import datetime as dt_type, date as date_type
    today_in_month = False
    for row in ws_month.iter_rows(min_row=2, values_only=False):
        val = row[0].value
        match = (
            (isinstance(val, (int, float)) and int(val) == today_serial) or
            (isinstance(val, dt_type) and val.date() == today) or
            (isinstance(val, date_type) and not isinstance(val, dt_type) and val == today)
        )
        if match:
            row[1].value = task_summary
            row[2].value = hours
            today_in_month = True
            break
    if not today_in_month:
        ws_month.append([today_serial, task_summary, hours])
        last = ws_month.max_row
        ws_month.cell(last, 1).number_format = "DD-MMM-YY"

    wb.save(LOCAL_TIMESHEET_PATH)
    print(f"  In-Out & {today.strftime('%b %y')} sheet updated for {today.strftime('%d-%b-%Y')}")


def update_tickets_sheet(tickets_worked_today: list = None):
    """
    For each open Jira ticket (or supplied list), pull changelog and update Tickets sheet.
    If tickets_worked_today is None, fetches all currently open tickets.
    Also fills in missing dates for existing tickets (backfill).
    """
    wb = load_workbook(LOCAL_TIMESHEET_PATH)
    ws = _get_tickets_sheet(wb)

    # Get current open tickets from Jira
    open_tickets = get_my_open_tickets()
    # Get resolved tickets since Dec 2025 (full history backfill)
    resolved_tickets = get_resolved_tickets_since("2025-12-01")

    all_jira_tickets = {t["key"]: t for t in open_tickets + resolved_tickets}
    ticket_keys_in_jira = set(all_jira_tickets.keys())

    # Also gather existing ticket keys from sheet to backfill
    existing_keys_in_sheet = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            existing_keys_in_sheet.add(str(row[1]).strip())

    # Union: open tickets + existing sheet tickets
    all_keys_to_process = ticket_keys_in_jira | existing_keys_in_sheet

    updated = 0
    added = 0

    for key in sorted(all_keys_to_process):
        try:
            data = get_ticket_changelog(key)
        except Exception as e:
            print(f"  Warning: Could not fetch {key}: {e}")
            continue

        fields = data.get("fields", {})
        status = fields.get("status", {}).get("name", "")
        due_raw = fields.get("duedate")
        created_raw = fields.get("created", "")[:10] if fields.get("created") else ""
        due_date = datetime.strptime(due_raw, "%Y-%m-%d").date() if due_raw else None
        created_date = datetime.strptime(created_raw, "%Y-%m-%d").date() if created_raw else None
        report = get_report_name(data)
        url = get_ticket_url(key)

        dates = extract_status_dates(
            data,
            ALL_STARTED_STATUSES,
            ALL_END_STATUSES,
            ALL_CLOSE_STATUSES
        )

        # Calculate derived fields
        started = dates["started_on"]
        end_d = dates["end_date"]
        close_d = dates["close_date"]
        today = date.today()

        days_open = None
        cycle_time = None
        if started:
            ref_end = close_d or end_d or today
            days_open = (ref_end - started).days
        if started and end_d:
            cycle_time = (end_d - started).days

        existing_row = _find_ticket_row(ws, key)

        if existing_row:
            # Update existing row — preserve Comments (col 12)
            ws.cell(existing_row, 3).value = status

            # "Only update if newer" guard for dates — prevents backward overwrites
            def _existing_date(col):
                val = ws.cell(existing_row, col).value
                if val is None:
                    return None
                if isinstance(val, (int, float)):
                    from datetime import timedelta
                    return date(1899, 12, 30) + timedelta(days=int(val))
                if hasattr(val, 'date'):
                    return val.date()
                return None

            # Started On: only write if not already set (first entry wins)
            if started and _existing_date(4) is None:
                _write_date_cell(ws.cell(existing_row, 4), started)

            # End Date: only write if Jira date is newer than what's in sheet
            if end_d:
                existing_end = _existing_date(5)
                if existing_end is None or end_d > existing_end:
                    _write_date_cell(ws.cell(existing_row, 5), end_d)

            # Close Date: only write if Jira date is newer
            if close_d:
                existing_close = _existing_date(6)
                if existing_close is None or close_d > existing_close:
                    _write_date_cell(ws.cell(existing_row, 6), close_d)

            if days_open is not None:
                ws.cell(existing_row, 7).value = days_open
            if cycle_time is not None:
                ws.cell(existing_row, 8).value = cycle_time
            if due_date:
                _write_date_cell(ws.cell(existing_row, 9), due_date)
            if created_date:
                _write_date_cell(ws.cell(existing_row, 10), created_date)
            if report and not ws.cell(existing_row, 11).value:
                ws.cell(existing_row, 11).value = report

            # Re-apply hyperlink to ticket key cell
            ws.cell(existing_row, 2).hyperlink = url
            ws.cell(existing_row, 2).font = Font(color="0563C1", underline="single")
            updated += 1

        else:
            # New row
            next_row = ws.max_row + 1
            sl = ws.max_row - 1  # Sl.No.

            ws.cell(next_row, 1).value = sl
            cell_key = ws.cell(next_row, 2)
            cell_key.value = key
            cell_key.hyperlink = url
            cell_key.font = Font(color="0563C1", underline="single")
            ws.cell(next_row, 3).value = status

            _write_date_cell(ws.cell(next_row, 4), started)
            _write_date_cell(ws.cell(next_row, 5), end_d)
            _write_date_cell(ws.cell(next_row, 6), close_d)
            if days_open is not None:
                ws.cell(next_row, 7).value = days_open
            if cycle_time is not None:
                ws.cell(next_row, 8).value = cycle_time
            _write_date_cell(ws.cell(next_row, 9), due_date)
            _write_date_cell(ws.cell(next_row, 10), created_date)
            ws.cell(next_row, 11).value = report
            added += 1

    # Set column widths
    col_widths = [6, 14, 18, 13, 13, 13, 10, 11, 13, 13, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(LOCAL_TIMESHEET_PATH)
    print(f"  Tickets sheet: {updated} updated, {added} new rows added.")


def export_monthly_report(month_name: str = None):
    """
    Export a manager-ready file with only In-Out + the specified month sheet.
    month_name: e.g. 'Apr 26'. Defaults to current month.
    Output: saved beside the main file as 'Timesheet_Apr26_Report.xlsx'
    """
    from openpyxl import Workbook
    import shutil
    import copy

    today = date.today()
    if month_name is None:
        month_name = today.strftime("%b %y")

    wb_src = load_workbook(LOCAL_TIMESHEET_PATH)

    if month_name not in wb_src.sheetnames:
        print(f"  Sheet '{month_name}' not found in timesheet.")
        return

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet in ["In-Out", month_name]:
        ws_src = wb_src[sheet]
        ws_dst = wb_out.create_sheet(sheet)
        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy.copy(cell.alignment)
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col_letter].width = dim.width

    folder = os.path.dirname(LOCAL_TIMESHEET_PATH)
    out_name = f"Timesheet_{month_name.replace(' ', '')}_{today.strftime('%Y%m%d')}.xlsx"
    out_path = os.path.join(folder, out_name)
    wb_out.save(out_path)
    print(f"  Monthly report saved: {out_path}")
    return out_path
