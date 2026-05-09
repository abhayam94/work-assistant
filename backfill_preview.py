"""
backfill_preview.py — Preview what Jira knows about existing tickets before updating.
Run this FIRST. Review output. Then confirm to proceed with eod_updater.py.

Shows a table: Ticket | Current Sheet Data | What Jira Has
"""

from colorama import init, Fore, Style
from openpyxl import load_workbook
from datetime import datetime, date
from jira_client import get_ticket_changelog, extract_status_dates, get_report_name
from config import LOCAL_TIMESHEET_PATH, ALL_STARTED_STATUSES, ALL_END_STATUSES, ALL_CLOSE_STATUSES

init(autoreset=True)
from logger import log_start, log_step, log_end


def _excel_to_date(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        from datetime import date as dt
        delta = int(val)
        return (dt(1899, 12, 30) + __import__('datetime').timedelta(days=delta)).strftime("%d-%b-%Y")
    if isinstance(val, (datetime, date)):
        return val.strftime("%d-%b-%Y")
    return str(val)


def run():
    print()
    print(Fore.CYAN + Style.BRIGHT + "  BACKFILL PREVIEW — Existing Tickets")
    print(Fore.WHITE + Style.DIM + "  Review this before running EOD updater for the first time.")
    print(Fore.WHITE + "─" * 90)

    wb = load_workbook(LOCAL_TIMESHEET_PATH, data_only=True)
    if "Tickets" not in wb.sheetnames:
        print(Fore.RED + "  No 'Tickets' sheet found in timesheet.")
        return

    ws = wb["Tickets"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        print(Fore.YELLOW + "  Tickets sheet is empty.")
        return

    print(f"\n  {'TICKET':<14} {'SHEET: Started':<16} {'JIRA: Started':<16} "
          f"{'JIRA: End':<14} {'JIRA: Close':<14} {'JIRA: Status':<22} {'JIRA: Report'}")
    print(Fore.WHITE + "─" * 115)

    for row in rows:
        ticket = str(row[1]).strip() if row[1] else ""
        if not ticket or ticket == "None":
            continue

        sheet_started = _excel_to_date(row[3])  # Col D

        try:
            data = get_ticket_changelog(ticket)
        except Exception as e:
            print(Fore.RED + f"  {ticket:<14} Could not fetch: {e}")
            continue

        status = data.get("fields", {}).get("status", {}).get("name", "")
        dates = extract_status_dates(data, ALL_STARTED_STATUSES, ALL_END_STATUSES, ALL_CLOSE_STATUSES)
        report = get_report_name(data)

        jira_started = dates["started_on"].strftime("%d-%b-%Y") if dates["started_on"] else Fore.RED + "Not found"
        jira_end = dates["end_date"].strftime("%d-%b-%Y") if dates["end_date"] else "—"
        jira_close = dates["close_date"].strftime("%d-%b-%Y") if dates["close_date"] else "—"

        match = "✓" if sheet_started and jira_started and sheet_started == jira_started else "≠"
        color = Fore.GREEN if match == "✓" else Fore.YELLOW

        print(
            f"  {Fore.WHITE}{ticket:<14}"
            f"{color}{str(sheet_started or '—'):<16}"
            f"{Fore.CYAN}{str(jira_started):<16}"
            f"{Fore.WHITE}{jira_end:<14}{jira_close:<14}"
            f"{Fore.YELLOW}{status:<22}"
            f"{Fore.WHITE}{report}"
        )

    print()
    print(Fore.WHITE + "─" * 90)
    log_end("backfill_preview", success=True)
    print(Fore.WHITE + Style.DIM + "\n  ≠ = mismatch between sheet and Jira  |  ✓ = match")
    print(Fore.YELLOW + "\n  Run eod_updater.py to apply updates. It will preserve your Comments column.\n")


if __name__ == "__main__":
    run()
