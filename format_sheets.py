"""
format_sheets.py — Apply consistent formatting to In-Out and Monthly sheets.
Also pre-populates the current (or next) month's dates upfront.

Run standalone:   python format_sheets.py
Also called from: timesheet_updater.py when a new month sheet is created.

Rules:
- In-Out:    weekdays only | alternate week row colors | bold border bottom of Friday
- Monthly:   all days | weekends marked bold | alternate week colors | bold border bottom of Sunday
- Both:      bold outer border around full table | light blue alternating week colors
"""

import calendar
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Style
from config import LOCAL_TIMESHEET_PATH

init(autoreset=True)
from logger import log_start, log_step, log_end, log_error

# --- COLORS ---
WEEK_COLOR_A    = "DEEAF1"   # Light blue
WEEK_COLOR_B    = "FFFFFF"   # White
WEEKEND_COLOR   = "F2F2F2"   # Light grey for weekend rows
HEADER_COLOR    = "2E75B6"   # Dark blue header
HEADER_FONT_CLR = "FFFFFF"   # White text on header

# --- BORDERS ---
THIN  = Side(style="thin")
THICK = Side(style="medium")
NO    = Side(style=None)

def _thick_bottom():
    return Border(left=NO, right=NO, top=NO, bottom=THICK)

def _thin_all():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _outer_border(ws, min_row, max_row, min_col, max_col):
    """Apply thick outer border around the full table."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            existing = cell.border
            left   = THICK if cell.column == min_col else existing.left
            right  = THICK if cell.column == max_col else existing.right
            top    = THICK if cell.row == min_row    else existing.top
            bottom = THICK if cell.row == max_row    else existing.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def _excel_to_date(val):
    """Convert Excel serial, datetime, or date to Python date."""
    from datetime import datetime as dt_type, date as date_type
    if isinstance(val, dt_type):
        return val.date()
    if isinstance(val, date_type):
        return val
    if isinstance(val, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(val))
    return None

def _excel_serial(d: date) -> int:
    return (d - date(1899, 12, 30)).days

def _header_style(ws, headers: list, col_widths: list):
    """Write and style header row."""
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True, color=HEADER_FONT_CLR)
        cell.fill = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 18


# ─────────────────────────────────────────────
# IN-OUT SHEET
# ─────────────────────────────────────────────

def format_inout_sheet(ws):
    """
    Format the In-Out sheet:
    - Header row styled
    - Weekdays only (no Sat/Sun)
    - Alternating week colors (week = Mon–Fri block)
    - Thick bottom border on Friday rows
    - Thick outer border around full table
    """
    headers = ["Date", "Task", "Hours", "Sign-In", "Sign-Out"]
    col_widths = [14, 45, 8, 10, 10]
    _header_style(ws, headers, col_widths)

    # Collect data rows with actual dates
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        d = _excel_to_date(row[0].value)
        if d:
            data_rows.append((d, row))

    if not data_rows:
        return

    # Assign week number (Mon=start) for color alternation
    # Week group: increment every Monday
    week_group = 0
    prev_monday = None

    for d, row in data_rows:
        monday = d - timedelta(days=d.weekday())  # Monday of this week
        if prev_monday is None or monday != prev_monday:
            week_group += 1
            prev_monday = monday

        fill_color = WEEK_COLOR_A if week_group % 2 == 1 else WEEK_COLOR_B
        fill = PatternFill("solid", start_color=fill_color)
        is_friday = (d.weekday() == 4)

        for cell in row:
            if cell.column > 5:
                continue
            cell.fill = fill
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="center")
            # Date column format
            if cell.column == 1:
                cell.number_format = "DD-MMM-YY"
                cell.value = _excel_serial(d)
            # Time columns
            if cell.column in (4, 5):
                cell.number_format = "HH:MM"

            # Border
            bottom = THICK if is_friday else THIN
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=bottom)

    # Outer border
    _outer_border(ws, 1, ws.max_row, 1, 5)


# ─────────────────────────────────────────────
# MONTHLY SHEET
# ─────────────────────────────────────────────

def format_monthly_sheet(ws, year: int, month: int):
    """
    Format a monthly sheet:
    - All calendar days included
    - Weekend rows: grey fill + bold 'Weekend' text
    - Alternating week colors for weekday rows
    - Thick bottom border on Sunday rows
    - Thick outer border around full table
    """
    headers = ["Date", "Task", "Hours"]
    col_widths = [14, 55, 8]
    _header_style(ws, headers, col_widths)

    # Read existing data into a dict: date → (task, hours)
    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _excel_to_date(row[0])
        if d:
            existing[d] = (row[1], row[2])

    # Clear data rows and rewrite all days of month
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font()

    # Generate all days in month
    _, days_in_month = calendar.monthrange(year, month)
    all_days = [date(year, month, d) for d in range(1, days_in_month + 1)]

    week_group = 0
    prev_monday = None

    for i, d in enumerate(all_days):
        row_num = i + 2
        is_weekend = d.weekday() >= 5  # Sat=5, Sun=6
        is_sunday  = d.weekday() == 6
        monday     = d - timedelta(days=d.weekday())

        if prev_monday is None or monday != prev_monday:
            week_group += 1
            prev_monday = monday

        # Fill color
        if is_weekend:
            fill = PatternFill("solid", start_color=WEEKEND_COLOR)
        else:
            fill_color = WEEK_COLOR_A if week_group % 2 == 1 else WEEK_COLOR_B
            fill = PatternFill("solid", start_color=fill_color)

        # Date cell
        date_cell = ws.cell(row=row_num, column=1)
        date_cell.value = _excel_serial(d)
        date_cell.number_format = "DD-MMM-YY"
        date_cell.fill = fill
        date_cell.alignment = Alignment(vertical="center")

        # Task cell
        task_cell = ws.cell(row=row_num, column=2)
        hours_cell = ws.cell(row=row_num, column=3)

        if is_weekend:
            task_cell.value = "Weekend"
            task_cell.font = Font(bold=True, color="888888")
            hours_cell.value = 0
        else:
            task_val, hours_val = existing.get(d, (None, None))
            task_cell.value = task_val
            task_cell.font = Font(name="Calibri", size=11)
            hours_cell.value = hours_val

        task_cell.fill = fill
        task_cell.alignment = Alignment(vertical="center")
        hours_cell.fill = fill
        hours_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Borders
        bottom = THICK if is_sunday else THIN
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = Border(
                left=THIN, right=THIN, top=THIN, bottom=bottom
            )

    # Outer border
    _outer_border(ws, 1, ws.max_row, 1, 3)


# ─────────────────────────────────────────────
# IN-OUT PRE-POPULATION
# ─────────────────────────────────────────────

def prepopulate_inout(ws, year: int, month: int):
    """Add all weekday rows for the month to In-Out sheet if not already present."""
    from datetime import datetime as dt_type, date as date_type

    # Collect existing dates
    existing_dates = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        d = _excel_to_date(row[0].value)
        if d:
            existing_dates.add(d)

    _, days_in_month = calendar.monthrange(year, month)
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if d.weekday() < 5 and d not in existing_dates:  # Mon–Fri only
            ws.append([_excel_serial(d), "", 0,
                       _excel_serial(date(1899, 12, 30) + timedelta(hours=11, minutes=30)),
                       _excel_serial(date(1899, 12, 30) + timedelta(hours=20, minutes=30))])
            last = ws.max_row
            ws.cell(last, 1).number_format = "DD-MMM-YY"
            ws.cell(last, 4).number_format = "HH:MM"
            ws.cell(last, 5).number_format = "HH:MM"

    # Sort rows by date after inserting
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data_rows.append(list(row))
    data_rows.sort(key=lambda r: r[0] if isinstance(r[0], (int, float)) else 0)

    # Rewrite sorted
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            ws.cell(i + 2, j + 1).value = val


# ─────────────────────────────────────────────
# MAIN ENTRY
# ─────────────────────────────────────────────

def apply_all_formatting(target_month: date = None):
    """
    Apply formatting to In-Out and all monthly sheets.
    Pre-populate current month if target_month provided.
    """
    wb = load_workbook(LOCAL_TIMESHEET_PATH)
    today = target_month or date.today()

    # Format In-Out
    log_start("format_sheets")
    if "In-Out" in wb.sheetnames:
        ws_io = wb["In-Out"]
        month_name = today.strftime("%b %y")
        prepopulate_inout(ws_io, today.year, today.month)
        format_inout_sheet(ws_io)
        log_step("In-Out", "OK", f"pre-populated {month_name}")
        print(Fore.GREEN + f"  ✓ In-Out sheet formatted & pre-populated for {month_name}")

    # Format all monthly sheets
    for sheet_name in wb.sheetnames:
        if sheet_name in ("In-Out", "Tickets"):
            continue
        try:
            # Parse month from sheet name e.g. 'May 26'
            parsed = date_from_sheet_name(sheet_name)
            if parsed:
                ws = wb[sheet_name]
                format_monthly_sheet(ws, parsed.year, parsed.month)
                log_step(sheet_name, "OK")
                print(Fore.GREEN + f"  ✓ {sheet_name} sheet formatted")
        except Exception as e:
            print(Fore.YELLOW + f"  ! Could not format '{sheet_name}': {e}")

    wb.save(LOCAL_TIMESHEET_PATH)
    log_end("format_sheets", success=True)
    print(Fore.CYAN + "\n  All sheets formatted and saved.\n")


def date_from_sheet_name(name: str):
    """Parse 'May 26' → date(2026, 5, 1). Returns None if unrecognised."""
    import re
    from datetime import datetime
    name = name.strip()
    # Match patterns like 'May 26', 'Dec 25', 'Jan 26'
    match = re.match(r"^([A-Za-z]{3})\s+(\d{2})$", name)
    if not match:
        return None
    try:
        return datetime.strptime(f"01 {match.group(1)} 20{match.group(2)}", "%d %b %Y").date()
    except Exception:
        return None


def prepopulate_next_month():
    """Pre-create and format next month's sheets in the last week of current month."""
    import calendar
    wb = load_workbook(LOCAL_TIMESHEET_PATH)
    today = date.today()
    next_month_first = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
    month_name = next_month_first.strftime("%b %y")

    # Monthly sheet
    if month_name not in wb.sheetnames:
        ws = wb.create_sheet(month_name)
        ws.append(["Date", "Task", "Hours"])
        format_monthly_sheet(ws, next_month_first.year, next_month_first.month)
        print(Fore.GREEN + f"  ✓ Created {month_name} monthly sheet")
    else:
        print(Fore.YELLOW + f"  {month_name} sheet already exists")

    # Pre-populate In-Out for next month weekdays
    if "In-Out" in wb.sheetnames:
        prepopulate_inout(wb["In-Out"], next_month_first.year, next_month_first.month)
        format_inout_sheet(wb["In-Out"])
        print(Fore.GREEN + f"  ✓ In-Out pre-populated for {month_name}")

    wb.save(LOCAL_TIMESHEET_PATH)
    print(Fore.CYAN + f"  Next month ({month_name}) is ready.\n")


if __name__ == "__main__":
    import sys
    print()
    print(Fore.CYAN + Style.BRIGHT + "  SHEET FORMATTER")
    print(Fore.WHITE + Style.DIM + "  Applying consistent formatting to all sheets")
    print(Fore.WHITE + "─" * 55)
    print()
    if "--next" in sys.argv:
        print(Fore.CYAN + "  Pre-creating next month sheets...\n")
        prepopulate_next_month()
    else:
        apply_all_formatting()
